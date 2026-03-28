import io
import msoffcrypto
import openpyxl

TEMPLATE = r"C:\Users\Usuario\Documents\ProyectoGrados\FGL 059 Monitoreo de Condiciones Ambientales del Laboratorio_V10 (7).xlsm"
PASSWORD = "Q1"

with open(TEMPLATE, "rb") as f:
    office_file = msoffcrypto.OfficeFile(f)
    office_file.load_key(password=PASSWORD)
    decrypted = io.BytesIO()
    office_file.decrypt(decrypted)

decrypted.seek(0)
wb = openpyxl.load_workbook(decrypted, keep_vba=True, data_only=False)

# Hojas de datos clave
for sheet_name in ['Datos - S1', 'Hoja1', 'Consolidado']:
    ws = wb[sheet_name]
    print(f"\n{'='*70}")
    print(f"HOJA: \"{sheet_name}\" — {ws.max_row} filas x {ws.max_column} cols")
    print(f"{'='*70}")

    for row in ws.iter_rows(min_row=1, max_row=30):
        row_data = []
        for cell in row:
            if cell.value is not None:
                val = str(cell.value)[:100]
                row_data.append(f"  {cell.coordinate}: {val}")
        if row_data:
            print(f"Fila {row[0].row}:")
            for item in row_data:
                print(item)

    print(f"\n--- Celdas combinadas ---")
    for m in list(ws.merged_cells.ranges)[:30]:
        print(f"  {m}")
