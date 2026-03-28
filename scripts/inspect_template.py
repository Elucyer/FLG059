import io
import msoffcrypto
import openpyxl

TEMPLATE = r"C:\Users\Usuario\Documents\ProyectoGrados\FGL 059 Monitoreo de Condiciones Ambientales del Laboratorio_V10 (7).xlsm"
PASSWORD = "Q1"

# Descifrar el archivo en memoria
with open(TEMPLATE, "rb") as f:
    office_file = msoffcrypto.OfficeFile(f)
    office_file.load_key(password=PASSWORD)
    decrypted = io.BytesIO()
    office_file.decrypt(decrypted)

decrypted.seek(0)
wb = openpyxl.load_workbook(decrypted, keep_vba=True, data_only=False)

print(f"\nHojas ({len(wb.sheetnames)}): {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"HOJA: \"{sheet_name}\" — {ws.max_row} filas x {ws.max_column} columnas")
    print(f"{'='*60}")

    printed = 0
    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            val = None
            if cell.value is not None:
                val = str(cell.value)[:80]
            if val:
                row_data.append(f"  {cell.coordinate}: {val}")
        if row_data:
            print(f"Fila {row[0].row}:")
            for item in row_data:
                print(item)
            printed += 1
        if printed >= 60:
            print(f"  ... ({ws.max_row - row[0].row} filas más)")
            break

    # Mostrar celdas combinadas
    if ws.merged_cells:
        merges = list(str(m) for m in ws.merged_cells.ranges)[:20]
        print(f"\nCeldas combinadas: {merges}")
